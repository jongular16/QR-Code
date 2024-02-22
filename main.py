from openpyxl import load_workbook
from email.message import EmailMessage
import smtplib
import ssl
import os

# To open the excel sheet 
book = load_workbook('GraduatedStudent.xlsx')
sheet = book.active

# To take information from the excel sheet
emails = []
outCount = 0
for row in sheet.iter_rows():
    outCount += 1
    count = 0
    if outCount != 1:
        for cell in row:
            count += 1
            if count == 2:
                emails.append(cell.value)

# To take the password
with open('The password.txt', 'r') as file:
    email_password = file.read()    

# To take the email sender
with open('Sender.txt', 'r') as file:
    email_sender = file.read()    

# To take the subject of the email
with open('Subject.txt', 'r') as file:
    subject = file.read()

# To take the body of the email
with open('body.txt', 'r') as file:
    body = file.read()

# Here we just sending emails
fPath = os.path.abspath('QR codes')
files = os.listdir(fPath)
con = ssl.create_default_context()
with smtplib.SMTP_SSL('smtp.gmail.com',465,context=con) as mt:
    mt.login(email_sender, email_password)    
    for i in range(len(emails)):
        recipient_email = emails[i]
        file = files[i]
        fll = r'{}\{}'.format(fPath,file)
        em = EmailMessage()
        em['From'] = email_sender
        em['Subject'] = subject
        em.set_content(body)
        with open(fll, 'rb') as attachment:
           em.add_attachment(attachment.read(), maintype='image', subtype='png', filename=fll)
        em['To'] = recipient_email
        mt.sendmail(email_sender, recipient_email, em.as_string())